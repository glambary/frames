import logging
from collections.abc import Generator
from types import TracebackType
from typing import Any

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from exception.custom import UnexpectedResultError
from services.base.service import BaseService
from services.excel.exceptions import ExcelNeverError, ExcelValueFromUserError
from services.excel.schemas.input import (
    CellCoordinatesInputSchema,
    RangeCellInputSchema,
    SheetInputSchema,
)
from services.excel.types import BaseModelChildType


class BaseExcelService(BaseService):
    """Базовый сервис для взаимодействия с excel файлами."""

    _workbook: Workbook

    def __init__(self, file: str) -> None:
        self._file = file

    def __enter__(self):
        self._workbook = load_workbook(filename=self._file, data_only=True)
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> None:
        if self._workbook:
            self._workbook.close()

    def _get_sheet(self, sheet_params: SheetInputSchema) -> Worksheet:
        """Возвращает страницу документа."""
        if not self._workbook:
            raise ExcelNeverError("Документ не открыт.")

        sheet_name = sheet_params.name
        sheet_index = sheet_params.index

        if sheet_name is not None:
            return self._workbook[sheet_name]

        if sheet_index is not None:
            return self._workbook.worksheets[sheet_index]

        raise ExcelNeverError(
            "При получении страницы, не передан sheet_name или sheet_index."
        )

    @staticmethod
    def _get_cell_value(
        sheet: Worksheet, cell: str | CellCoordinatesInputSchema
    ) -> Any:
        if isinstance(cell, str):
            return sheet[cell].value
        if isinstance(cell, CellCoordinatesInputSchema):
            return sheet.cell(row=cell.row, column=cell.column).value

        raise ExcelNeverError(
            f"Некорректный тип ячейки: {type(cell).__name__}"
        )

    @staticmethod
    def _get_iter_rows(
        sheet: Worksheet, range_cell_params: RangeCellInputSchema
    ) -> Generator[tuple[Any, ...], None, None]:
        iter_rows = sheet.iter_rows(
            min_row=range_cell_params.start_row,
            max_row=range_cell_params.end_row,
            min_col=range_cell_params.start_column,
            max_col=range_cell_params.end_column,
            values_only=True,
        )
        return iter_rows


class ExcelService(BaseExcelService):
    """Сервис для взаимодействия с excel файлами."""

    BaseReturnType = tuple[Any, ...] | BaseModelChildType

    def get_rows_data(
        self,
        sheet_params: SheetInputSchema,
        range_cell_params: RangeCellInputSchema,
        columns: dict[int, str] | None = None,
        validate_to_schema: type[BaseModelChildType] | None = None,
    ) -> tuple[tuple[BaseReturnType, ...], dict[int, ValidationError] | None]:
        """Возвращает данные из страницы построчно."""
        common_log_information = self._get_common_log_information(
            self.get_rows_data
        )

        sheet: Worksheet = self._get_sheet(sheet_params)

        iter_rows = self._get_iter_rows(sheet, range_cell_params)

        if not validate_to_schema:
            return tuple(iter_rows), None

        # Проверка, что для ответа в pydantic схемах переданы
        # необходимые параметры
        for a in (columns, validate_to_schema):
            if not a:
                raise UnexpectedResultError(
                    common_log_information
                    + f"Для получения ответа в pydantic схемах, "
                    f"{a} обязателен."
                )

        result = []
        rows_with_exc = {}
        for n, row in enumerate(iter_rows, start=1):
            data = {
                field_name: row[column - 1]
                for column, field_name in columns.items()
            }

            # если ошибка валидации, то пропускаем строку
            try:
                result.append(validate_to_schema.model_validate(data))
            except ValidationError as exc:
                if len({r for r in data.values() if r is not None}) > 1:
                    rows_with_exc[n] = exc
                    logging.warning(
                        common_log_information
                        + f"Значения строки {n}-{list(data.values())} заполнены некорректно.",
                        exc_info=exc,
                    )
                continue

        return tuple(result), rows_with_exc

    def get_cell_values(
        self,
        sheet_params: SheetInputSchema,
        cells: (
            tuple[str | CellCoordinatesInputSchema, ...]
            | dict[str, str | CellCoordinatesInputSchema]
        ),
        validate_to_schema: type[BaseModelChildType] | None = None,
    ) -> BaseReturnType:
        """Возвращает значения из указанных ячеек."""
        common_log_information = self._get_common_log_information(
            self.get_cell_values
        )

        if not cells:
            return ()

        # Проверка формата данных и validate_to_schema
        if isinstance(cells, tuple) and validate_to_schema:
            raise ExcelNeverError(
                common_log_information + "Для `cells` в формате tuple,"
                "`validate_to_schema` должен быть None."
            )
        if isinstance(cells, dict) and not validate_to_schema:
            raise ExcelNeverError(
                common_log_information + "Для `cells` в формате dict,"
                "`validate_to_schema` должен быть передан."
            )

        sheet: Worksheet = self._get_sheet(sheet_params)

        # Обработка данных
        if isinstance(cells, tuple):
            # Сырые данные (только tuple)
            return tuple(self._get_cell_value(sheet, cell) for cell in cells)

        if isinstance(cells, dict):
            # Преобразование данных для схемы
            data_to_validate = {
                field_name: self._get_cell_value(sheet, cell)
                for field_name, cell in cells.items()
            }
            try:
                schema = validate_to_schema.model_validate(data_to_validate)
            except ValidationError as exc:
                msg = f"Ошибка при валидации схемы {validate_to_schema}."
                logging.warning(
                    common_log_information + msg,
                    exc_info=exc,
                )
                raise ExcelValueFromUserError(
                    common_log_information + msg
                ) from exc

            return schema

        # Если формат cells не распознан
        raise ExcelNeverError(
            common_log_information
            + f"Некорректный формат cells: {type(cells).__name__}"
        )
