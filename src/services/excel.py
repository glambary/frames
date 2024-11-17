import logging
from typing import Any, assert_never

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from exception.custom import UnexpectedResultError
from schemas.excel.input import (
    CellCoordinatesInputSchema,
    RangeCellInputSchema,
    SheetInputSchema,
)
from schemas.types import BaseModelType


class ExcelService:
    """Сервис для взаимодействия с файлами excel."""

    BaseReturnType = tuple[(tuple[Any, ...] | BaseModelType), ...]

    def __init__(self, file: str) -> None:
        self.workbook: Workbook = load_workbook(filename=file)

    def _get_sheet(self, sheet_params: SheetInputSchema) -> Worksheet:
        """Возвращает страницу документа."""
        sheet_name = sheet_params.name
        sheet_index = sheet_params.index

        if sheet_name is not None:
            return self.workbook[sheet_name]

        if sheet_index is not None:
            return self.workbook.worksheets[sheet_index]

        raise UnexpectedResultError("Недостижимый участок кода.")

    def get_rows_data(
        self,
        sheet_params: SheetInputSchema,
        range_cell_params: RangeCellInputSchema,
        raw_answer: bool = True,
        columns: dict[int, str] | None = None,
        validate_to_schema: type[BaseModelType] | None = None,
    ) -> BaseReturnType | tuple[BaseReturnType, dict[int, ValidationError]]:
        """
        Возвращает информацию из excel страницы построчно.

        :param sheet_params: Параметры для выбора страницы;
        :param range_cell_params: Параметры для выбора строк, столбцов;
        :param raw_answer: Вернуть сырой ответ или нет;
        :param columns: Если ответ НЕ сырой,
            то формат {номер_колонки: имя_поля из validate_to_schema}
        :param validate_to_schema: Pydantic схема;

        Если схема validate_to_schema не сможет создаться, то строка будет
        пропущена.
        """
        sheet: Worksheet = self._get_sheet(sheet_params)

        iter_rows = sheet.iter_rows(
            min_row=range_cell_params.start_row,
            max_row=range_cell_params.end_row,
            min_col=range_cell_params.start_column,
            max_col=range_cell_params.end_column,
            values_only=True,
        )

        if raw_answer:
            return tuple(iter_rows)

        for a in (columns, validate_to_schema):
            if not a:
                raise ValueError(
                    f"Для получения ответа в pydantic схемах, {a} обязателен."
                )

        result = []
        rows_with_exc = {}
        for n, row in enumerate(iter_rows, start=1):
            data = {
                field_name: row[column]
                for column, field_name in columns.items()
            }

            # если ошибка валидации, то пропускаем строку
            try:
                result.append(validate_to_schema.model_validate(data))
            except ValidationError as exc:
                rows_with_exc[n] = exc
                # logging.warning(
                #     f"Значения строки {n} заполнены не полностью "
                #     f"или некорректно.",
                #     exc_info=exc,
                # )
                continue

        return tuple(result), rows_with_exc

    def get_cell_values(
        self,
        sheet_params: SheetInputSchema,
        cells: tuple[str | CellCoordinatesInputSchema, ...],
        raw_answer: bool = True,
        name_fields: tuple[str, ...] | None = None,
        validate_to_schema: type[BaseModelType] | None = None,
    ) -> BaseReturnType:
        """
        Возвращает значения из указанных ячеек.

        :param sheet_params: Параметры для выбора страницы;
        :param cells: Кортеж адресов ячеек (например, "A1", "B2")
            или координат CellCoordinatesSchema.
        :param raw_answer: Вернуть сырой ответ или нет;
        :param name_fields: Если ответ НЕ сырой, то поле обязательное.
            Внимание len(cells) и len(name_fields) должно совпадать.
        :param validate_to_schema: Pydantic схема;
        """
        sheet: Worksheet = self._get_sheet(sheet_params)

        if not cells:
            return ()

        match cells[0]:
            case str():
                func_get_cell = lambda cell: sheet[cell].value  # noqa: E731
            case CellCoordinatesInputSchema():
                func_get_cell = lambda cell: sheet.cell(  # noqa: E731
                    row=cell.row, column=cell.column
                ).value
            case _ as unexpected_result:
                assert_never(unexpected_result)

        values = tuple(func_get_cell(cell) for cell in cells)

        if raw_answer:
            return values

        if not name_fields or len(name_fields) != len(cells):
            raise ValueError(
                "Некорректные данные name_fields,"
                "для получения ответа в pydantic схемах."
            )

        return validate_to_schema.model_validate(
            dict(zip(name_fields, values, strict=True))
        )
